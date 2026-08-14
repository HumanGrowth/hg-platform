"""Capa Empresa · TASK 4: bulk import de miembros desde Excel.

Cubre: archivo válido mixto (altas + updates + errores), idempotencia por email,
exceso de licencias, organización inexistente, manager inexistente, y la descarga
de la plantilla.
"""
from __future__ import annotations

from io import BytesIO
from uuid import uuid4

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import func, select

from hg.modules.company.bulk_service import TEMPLATE_HEADERS, XLSX_MIME
from hg.modules.identity.invitations import Invitation
from hg.modules.identity.models import User, UserRole

API = "/api/v1"


def _xlsx(rows: list[list[str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(TEMPLATE_HEADERS)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _post(client: TestClient, headers: dict[str, str], content: bytes):
    return client.post(
        f"{API}/company/members/bulk-import",
        headers=headers,
        files={"file": ("miembros.xlsx", content, XLSX_MIME)},
    )


def _email() -> str:
    return f"u-{uuid4().hex[:10]}@hgtest.test"


def _setup(factory, *, company=None, licenses_total=None):
    """Crea org (slug único) + un company_admin. Devuelve (slug, org, admin)."""
    slug = f"co-{uuid4().hex[:8]}"
    org = factory.make_org(slug=slug, company=company, licenses_total=licenses_total)
    admin = factory.make_user(org=org, role=UserRole.company_admin)
    return slug, org, admin


# ─────────────────────────── Feliz + mixto ───────────────────────────


def test_bulk_mixed_creates_updates_and_reports_errors(
    client: TestClient, factory, auth_headers
) -> None:
    slug, org, admin = _setup(factory, licenses_total=None)  # sin cap propio de org
    exist_email = _email()
    new_email = _email()
    existing = factory.make_user(
        org=org, email=exist_email, full_name="Nombre Viejo", role=UserRole.collaborator
    )

    content = _xlsx(
        [
            [slug, exist_email, "Nombre Nuevo", "manager", ""],  # update
            [slug, new_email, "Persona Nueva", "collaborator", ""],  # alta
            ["no-existe", _email(), "Fulano", "collaborator", ""],  # org mala
            [slug, "sin-arroba", "Malo", "collaborator", ""],  # email inválido
        ]
    )
    res = _post(client, auth_headers(admin), content)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["total"] == 4
    assert body["creados"] == 1
    assert body["actualizados"] == 1
    assert body["errores"] == 2

    by_email = {r["email"]: r for r in body["filas"]}
    assert by_email[exist_email]["estado"] == "actualizado"
    assert by_email[new_email]["estado"] == "creado"

    s = factory.session
    s.expire_all()
    updated = s.get(User, existing.id)
    assert updated.full_name == "Nombre Nuevo"
    assert updated.role == UserRole.manager
    inv = s.scalar(select(Invitation).where(func.lower(Invitation.email) == new_email))
    assert inv is not None and inv.accepted_at is None


def test_bulk_is_idempotent_by_email(client: TestClient, factory, auth_headers) -> None:
    """Re-subir el mismo email no duplica: la 2da corrida reporta 'actualizado'
    (invitación ya pendiente) y no crea una segunda invitación."""
    slug, _org, admin = _setup(factory, licenses_total=None)
    email = _email()
    content = _xlsx([[slug, email, "Dup", "collaborator", ""]])

    first = _post(client, auth_headers(admin), content)
    assert first.json()["creados"] == 1

    second = _post(client, auth_headers(admin), content)
    assert second.json()["actualizados"] == 1
    assert second.json()["creados"] == 0

    s = factory.session
    count = s.scalar(
        select(func.count()).select_from(Invitation).where(
            func.lower(Invitation.email) == email
        )
    )
    assert count == 1


def test_bulk_duplicate_email_in_file_errors(client: TestClient, factory, auth_headers) -> None:
    slug, _org, admin = _setup(factory, licenses_total=None)
    email = _email()
    content = _xlsx(
        [
            [slug, email, "Uno", "collaborator", ""],
            [slug, email, "Dos", "collaborator", ""],
        ]
    )
    body = _post(client, auth_headers(admin), content).json()
    assert body["creados"] == 1
    assert body["errores"] == 1
    err = next(r for r in body["filas"] if r["estado"] == "error")
    assert "duplicado" in err["motivo"].lower()


# ─────────────────────────── Límites y validaciones ───────────────────────────


def test_bulk_respects_license_pool(client: TestClient, factory, auth_headers) -> None:
    """Con pool=2 y el admin ya activo (1 asiento), solo entra 1 alta más; la
    siguiente queda en error por exceso de pool."""
    company = factory.make_company(licenses_total=2)
    slug, _org, admin = _setup(factory, company=company, licenses_total=None)  # admin consume 1/2

    content = _xlsx(
        [
            [slug, _email(), "A", "collaborator", ""],
            [slug, _email(), "B", "collaborator", ""],
        ]
    )
    body = _post(client, auth_headers(admin), content).json()
    assert body["creados"] == 1
    assert body["errores"] == 1
    err = next(r for r in body["filas"] if r["estado"] == "error")
    assert "pool" in err["motivo"].lower()


def test_bulk_manager_must_exist_and_be_manager(
    client: TestClient, factory, auth_headers
) -> None:
    slug, org, admin = _setup(factory, licenses_total=None)
    notmgr_email = _email()
    factory.make_user(org=org, email=notmgr_email, role=UserRole.collaborator)

    content = _xlsx(
        [
            [slug, _email(), "X", "collaborator", _email()],  # manager inexistente
            [slug, _email(), "Y", "collaborator", notmgr_email],  # existe pero no es manager
        ]
    )
    body = _post(client, auth_headers(admin), content).json()
    assert body["errores"] == 2
    motivos = " ".join(r["motivo"].lower() for r in body["filas"])
    assert "manager_email no existe" in motivos
    assert "no es manager" in motivos


def test_bulk_bad_file_returns_400(client: TestClient, factory, auth_headers) -> None:
    _slug, _org, admin = _setup(factory, licenses_total=None)
    res = _post(client, auth_headers(admin), b"esto no es un xlsx")
    assert res.status_code == 400, res.text


def test_bulk_missing_column_returns_400(client: TestClient, factory, auth_headers) -> None:
    _slug, _org, admin = _setup(factory, licenses_total=None)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["email", "nombre_completo"])  # faltan columnas
    ws.append([_email(), "A"])
    buf = BytesIO()
    wb.save(buf)
    res = _post(client, auth_headers(admin), buf.getvalue())
    assert res.status_code == 400, res.text
    assert "columna" in res.json()["detail"].lower()


# ─────────────────────────── Template ───────────────────────────


def test_bulk_template_download(client: TestClient, factory, auth_headers) -> None:
    _slug, _org, admin = _setup(factory, licenses_total=None)
    res = client.get(
        f"{API}/company/members/bulk-import/template", headers=auth_headers(admin)
    )
    assert res.status_code == 200, res.text
    assert "spreadsheet" in res.headers["content-type"]
    wb = openpyxl.load_workbook(BytesIO(res.content))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header == TEMPLATE_HEADERS
