"""Bulk import de miembros desde Excel (Capa Empresa · TASK 4).

Sube un `.xlsx` con altas/updates de colaboradores de una Empresa. Es
**idempotente por email dentro de la Empresa** (crea invitación o actualiza el
user existente, nunca duplica) y **tolerante a filas malas** (jamás falla todo
por una fila: cada fila devuelve `creado | actualizado | error` con motivo).

Columnas (ver `TEMPLATE_HEADERS`): ``organizacion`` (slug o nombre exacto de una
org de la Empresa), ``email``, ``nombre_completo``, ``rol``
(``manager`` | ``collaborator``), ``manager_email`` (opcional: enlaza el
colaborador a un manager existente de esa org).

Reglas de licencia: se proyecta el uso del **pool de la Empresa** y del **cap de
la org** a lo largo del archivo, de modo que un Excel no pueda pasarse del cupo
(las filas que excederían quedan en error, las anteriores se aplican).

Aislamiento de fila: cada fila corre en un SAVEPOINT (`begin_nested`) — si su
escritura falla, se revierte solo esa fila sin envenenar la transacción.
"""
from __future__ import annotations

import re
from collections.abc import Callable
from dataclasses import dataclass
from io import BytesIO
from uuid import UUID

import openpyxl
from fastapi import HTTPException
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from hg.modules.company import service as company_service
from hg.modules.identity import service as identity_service
from hg.modules.identity.invitations import Invitation
from hg.modules.identity.models import Organization, User, UserRole

TEMPLATE_HEADERS = ["organizacion", "email", "nombre_completo", "rol", "manager_email"]
# MIME oficial del .xlsx (para el Content-Type de la plantilla y el upload).
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ROLE_ALIASES: dict[str, UserRole] = {
    "manager": UserRole.manager,
    "collaborator": UserRole.collaborator,
    "colaborador": UserRole.collaborator,  # alias ES tolerado
}
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_MAX_ROWS = 5000


@dataclass
class RowResult:
    fila: int
    email: str
    estado: str  # creado | actualizado | error
    motivo: str | None = None


class BulkImportError(Exception):
    """Error de archivo completo (headers/formato) → 400, no un error por fila."""


class _RowError(Exception):
    """Error de una fila puntual → se reporta en su RowResult, no aborta el lote."""


# ─────────────────────────── Parseo del .xlsx ───────────────────────────


def _parse_rows(content: bytes) -> list[dict[str, str]]:
    """Lee el workbook y devuelve una fila-dict por registro (headers tolerantes a
    orden/mayúsculas). Levanta ``BulkImportError`` si el archivo o los headers no
    sirven."""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl levanta variados; homogeneizamos
        raise BulkImportError(f"no se pudo leer el .xlsx: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise BulkImportError("el archivo no tiene hojas")

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        raise BulkImportError("el archivo está vacío")

    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    idx: dict[str, int] = {}
    for col in TEMPLATE_HEADERS:
        if col not in norm:
            raise BulkImportError(f"falta la columna obligatoria '{col}'")
        idx[col] = norm.index(col)

    records: list[dict[str, str]] = []
    for n, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # fila en blanco → se ignora
        rec = {
            col: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
            for col, i in idx.items()
        }
        rec["_fila"] = str(n)
        records.append(rec)
        if len(records) > _MAX_ROWS:
            raise BulkImportError(f"el archivo excede el máximo de {_MAX_ROWS} filas")
    return records


# ─────────────────────────── Import ───────────────────────────


def bulk_import(
    db: Session, *, company_id: UUID, actor: User, content: bytes
) -> list[RowResult]:
    """Procesa el `.xlsx` fila por fila y devuelve el reporte. Cada fila es un
    SAVEPOINT independiente; el commit final lo hace la dependencia del router."""
    records = _parse_rows(content)
    company = company_service._get_company(db, company_id)

    orgs = list(
        db.scalars(select(Organization).where(Organization.company_id == company_id)).all()
    )
    by_slug = {o.slug.strip().lower(): o for o in orgs}
    by_name = {o.name.strip().lower(): o for o in orgs}

    # Proyección de licencias a lo largo del archivo (pool empresa + cap org).
    pool_total = company.licenses_total
    pool_used = identity_service.company_active_users(db, company_id)
    org_used: dict[UUID, int] = {o.id: (o.licenses_used or 0) for o in orgs}

    results: list[RowResult] = []
    seen: set[str] = set()

    for rec in records:
        fila = int(rec["_fila"])
        email = rec["email"].strip().lower()
        if not email:
            results.append(RowResult(fila, "", "error", "email vacío"))
            continue
        if email in seen:
            results.append(RowResult(fila, email, "error", "email duplicado en el archivo"))
            continue
        seen.add(email)

        try:
            with db.begin_nested():
                estado, seat_org = _process_row(
                    db, company_id=company_id, actor=actor, rec=rec, email=email,
                    by_slug=by_slug, by_name=by_name,
                    pool_total=pool_total, pool_used=pool_used, org_used=org_used,
                )
        except _RowError as exc:
            results.append(RowResult(fila, email, "error", str(exc)))
            continue
        except HTTPException as exc:  # p.ej. check de licencias del flujo de identity
            results.append(RowResult(fila, email, "error", str(exc.detail)))
            continue

        if seat_org is not None:  # consumió un asiento → actualizar la proyección
            pool_used += 1
            org_used[seat_org] = org_used.get(seat_org, 0) + 1
        results.append(RowResult(fila, email, estado))

    return results


def _process_row(
    db: Session, *, company_id: UUID, actor: User, rec: dict[str, str], email: str,
    by_slug: dict[str, Organization], by_name: dict[str, Organization],
    pool_total: int, pool_used: int, org_used: dict[UUID, int],
) -> tuple[str, UUID | None]:
    """Valida + aplica una fila. Devuelve ``(estado, org_que_consumió_asiento)``;
    el segundo es ``None`` si la fila no consumió licencia (update sin reactivar o
    invitación pendiente ya existente)."""
    if not _EMAIL_RE.match(email):
        raise _RowError("email inválido")

    org = _resolve_org(rec["organizacion"], by_slug, by_name)
    role = _resolve_role(rec["rol"])
    full_name = rec["nombre_completo"].strip()

    manager = None
    if rec["manager_email"].strip():
        manager = _resolve_manager(db, org.id, rec["manager_email"].strip().lower())

    def reserve() -> None:
        """Chequea pool empresa + cap org contra la proyección; levanta si excede."""
        if pool_used >= pool_total:
            raise _RowError("se excede el pool de licencias de la empresa")
        if org.licenses_total is not None and org_used.get(org.id, 0) >= org.licenses_total:
            raise _RowError("se excede el cupo de licencias de la organización")

    existing = db.scalar(
        select(User).where(User.company_id == company_id, func.lower(User.email) == email)
    )
    if existing is not None:
        return _update_existing(existing, org, role, full_name, manager, reserve)

    # ¿Ya hay una invitación pendiente para este email en alguna org de la Empresa?
    pending = db.scalar(
        select(Invitation)
        .join(Organization, Organization.id == Invitation.org_id)
        .where(
            Organization.company_id == company_id,
            func.lower(Invitation.email) == email,
            Invitation.accepted_at.is_(None),
            Invitation.revoked_at.is_(None),
        )
    )
    if pending is not None:
        return "actualizado", None  # ya invitado → no se duplica

    reserve()
    identity_service.create_invitation(
        db, org_id=org.id, email=rec["email"].strip(), role=role, invited_by=actor,
        name=full_name or None,
    )
    return "creado", org.id


def _update_existing(
    user: User, org: Organization, role: UserRole, full_name: str,
    manager: User | None, reserve: Callable[[], None],
) -> tuple[str, UUID | None]:
    """Aplica los cambios de la fila a un user existente. Reactiva si estaba
    inactivo (consume asiento). Devuelve ``("actualizado", org_del_asiento|None)``."""
    if full_name:
        user.full_name = full_name
    if user.role != role:
        user.role = role
    if org.id != user.org_id:
        user.org_id = org.id
        user.company_id = org.company_id
        user.manager_id = None  # el manager viejo era de otra org
    if manager is not None:
        if manager.org_id != user.org_id:
            raise _RowError("el manager pertenece a otra organización")
        user.manager_id = manager.id
    seat_org: UUID | None = None
    if not user.is_active:
        reserve()
        user.is_active = True
        org.licenses_used = (org.licenses_used or 0) + 1
        seat_org = org.id
    return "actualizado", seat_org


def _resolve_org(
    raw: str, by_slug: dict[str, Organization], by_name: dict[str, Organization]
) -> Organization:
    key = raw.strip().lower()
    if not key:
        raise _RowError("falta la organización")
    org = by_slug.get(key) or by_name.get(key)
    if org is None:
        raise _RowError(f"la organización '{raw}' no pertenece a la empresa")
    return org


def _resolve_role(raw: str) -> UserRole:
    role = _ROLE_ALIASES.get(raw.strip().lower())
    if role is None:
        raise _RowError(f"rol inválido '{raw}' (usar manager o collaborator)")
    return role


def _resolve_manager(db: Session, org_id: UUID, manager_email: str) -> User:
    manager = db.scalar(
        select(User).where(User.org_id == org_id, func.lower(User.email) == manager_email)
    )
    if manager is None:
        raise _RowError("manager_email no existe en la organización")
    if manager.role != UserRole.manager:
        raise _RowError("manager_email no es manager de la organización")
    return manager


# ─────────────────────────── Template ───────────────────────────


def build_template_xlsx() -> bytes:
    """Genera el `.xlsx` de plantilla: headers + una fila de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Miembros"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["mi-org", "persona@empresa.com", "Nombre Apellido", "collaborator", ""])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
